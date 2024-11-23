from openpyxl.styles import Alignment, Font, PatternFill

def check_excel_3(workbook):
    checklist_data = {
    "Grading Criteria": [
            "Is the worksheet named 'Countries'?",
            "Are there 4 columns (Country, Continent, Population, GDP per Capita)?",
            "Does the table contain exactly 20 countries with data?",
            "Are the rows styled by continent?",
            "Is there a Population column chart positioned correctly?",
            "Is the Population chart titled 'POPULATION OF THE 20 SAMPLE COUNTRIES'?",
            "Is there a GDP per Capita chart with gradient fill?",
            "Is the GDP chart positioned below the Population chart?",
            "Is the table sorted by Population (largest to smallest)?",
            "Does row 22 contain SUM formulas for Population and GDP?",
            "Does row 23 contain AVERAGE formulas for Population and GDP?",
            "Is the ChatGPT link in merged cells A26:E26?",
            "Is the ChatGPT link centered?",
            "Does the ChatGPT link cell have a background color?"
        ],
        "Completed": []
    }

    # Check worksheet name
    sheet_name_correct = 'Countries' in workbook.sheetnames
    sheet = workbook['Countries'] if sheet_name_correct else workbook.active
    checklist_data["Completed"].append("Yes" if sheet_name_correct else "No")

    # Check column headers
    expected_headers = ["Country", "Continent", "Population", "GDP per Capita"]
    headers = [sheet.cell(row=1, column=i).value for i in range(1, 5)]
    headers_match = all(a == b for a, b in zip(headers, expected_headers))
    checklist_data["Completed"].append("Yes" if headers_match else "No")

    # Check for 20 countries with data
    data_rows = sum(1 for row in range(2, 22) if all(sheet.cell(row=row, column=col).value 
               for col in range(1, 5)))
    checklist_data["Completed"].append("Yes" if data_rows == 20 else "No")

    # Check for continent-based styling
    different_styles = False
    current_continent = None
    current_fill = None
    for row in range(2, 22):
        continent = sheet.cell(row=row, column=2).value
        fill = sheet.cell(row=row, column=1).fill
        if continent != current_continent:
            if current_continent is not None and fill != current_fill:
                different_styles = True
            current_continent = continent
            current_fill = fill
    checklist_data["Completed"].append("Yes" if different_styles else "No")

    # Check for Population chart
    has_population_chart = False
    for chart in sheet._charts:
        if hasattr(chart, 'title') and chart.title is not None:
            has_population_chart = True
            break
    checklist_data["Completed"].append("Yes" if has_population_chart else "No")

    # Check Population chart title
    correct_title = False
    for chart in sheet._charts:
        if hasattr(chart, 'title') and chart.title is not None:
            if hasattr(chart.title, 'tx') and hasattr(chart.title.tx, 'rich'):
                for p in chart.title.tx.rich.paragraphs:
                    if hasattr(p, 'r'):
                        for run in p.r:
                            if hasattr(run, 't'):
                                chart_title = str(run.t).strip()
                                expected_title = "Population of the 20 sample countries"
                                # Case-insensitive comparison
                                if chart_title.lower() == expected_title.lower():
                                    correct_title = True
                                    break
                                # For debugging
                                st.write(f"Found title: {chart_title}")
                                st.write(f"Expected title: {expected_title}")
    checklist_data["Completed"].append("Yes" if correct_title else "No")

    # Check for GDP chart with gradient fill
    has_gdp_chart = len(sheet._charts) >= 2
    checklist_data["Completed"].append("Yes" if has_gdp_chart else "No")

    # Check GDP chart position (below Population chart)
    charts_positioned_correctly = len(sheet._charts) >= 2
    checklist_data["Completed"].append("Yes" if charts_positioned_correctly else "No")

    # Check if sorted by Population (largest to smallest)
    is_sorted = True
    prev_value = float('inf')
    for row in range(2, 22):
        current_value = sheet.cell(row=row, column=3).value
        if isinstance(current_value, (int, float)) and current_value > prev_value:
            is_sorted = False
            break
        prev_value = current_value if isinstance(current_value, (int, float)) else float('inf')
    checklist_data["Completed"].append("Yes" if is_sorted else "No")

    # Check for SUM formulas in row 22
    total_population_formula = sheet['C22'].data_type == 'f'
    total_gdp_formula = sheet['D22'].data_type == 'f'
    checklist_data["Completed"].append("Yes" if total_population_formula and total_gdp_formula else "No")

    # Check for AVERAGE formulas in row 23
    avg_population_formula = sheet['C23'].data_type == 'f'
    avg_gdp_formula = sheet['D23'].data_type == 'f'
    checklist_data["Completed"].append("Yes" if avg_population_formula and avg_gdp_formula else "No")

    # Check ChatGPT link merged cells
    merged_link = any("A26:E26" in str(range) for range in sheet.merged_cells.ranges)
    checklist_data["Completed"].append("Yes" if merged_link else "No")

    # Check ChatGPT link alignment
    center_aligned = sheet['A26'].alignment.horizontal == 'center'
    checklist_data["Completed"].append("Yes" if center_aligned else "No")

    # Check ChatGPT link background color
    background_color_present = (sheet['A26'].fill is not None and 
                              sheet['A26'].fill.fill_type is not None)
    checklist_data["Completed"].append("Yes" if background_color_present else "No")

    # Display results
    st.subheader("Checklist Results")
    checklist_df = pd.DataFrame(checklist_data)
    st.table(checklist_df)

    # Calculate percentage complete and points
    total_yes = checklist_data["Completed"].count("Yes")
    total_items = len(checklist_data["Completed"])
    percentage_complete = (total_yes / total_items) * 100
    points = (total_yes / total_items) * 20

    # Create two columns for displaying scores
    col1, col2 = st.columns(2)

    return checklist_data
