from openpyxl.styles import Alignment, Font, PatternFill

def check_excel_1(workbook):
    sheet = workbook.active
    checklist_data = {
        "Grading Criteria": [
            "Does the dataset have exactly 7 columns (A-G)?",
            "Are there exactly 10 rows of data in the dataset?",
            "Are the first 6 column headers named ID, First Name, Last Name, Date of Birth, Hometown, Occupation?",
            "Does the 7th column have a meaningful header and consistent data?",
            "Are styles applied (e.g., alternating row colors or banded table)?",
            "Are the headers in row 1 bolded with a background color?",
            "Are the columns properly aligned (text, dates, and numbers)?",
            "Is the ChatGPT hyperlink centered in row 13?",
            "Are cells A13:G13 merged in row 13?",
            "Does row 13 have a background color and meaningful hyperlink text?"
        ],
        "Completed": []
    }

    # 1. Check if there are exactly 7 columns
    num_columns = sheet.max_column
    checklist_data["Completed"].append("Yes" if num_columns == 7 else "No")

    # 2. Check if there are exactly 10 rows of data
    data_rows = sum(1 for row in range(2, 12) if any(sheet.cell(row=row, column=col).value for col in range(1, 8)))
    checklist_data["Completed"].append("Yes" if data_rows == 10 else "No")

    # 3. Check first 6 column headers
    expected_headers = ["ID", "First Name", "Last Name", "Date of Birth", "Hometown", "Occupation"]
    headers = [sheet.cell(row=1, column=i).value for i in range(1, 7)]
    headers_match = all(a == b for a, b in zip(headers, expected_headers))
    checklist_data["Completed"].append("Yes" if headers_match else "No")

    # 4. Check 7th column header and data
    last_col_header = sheet.cell(row=1, column=7).value
    last_col_has_data = all(sheet.cell(row=i, column=7).value for i in range(2, 12))
    checklist_data["Completed"].append("Yes" if last_col_header and last_col_has_data else "No")

    # 5. Check if styles are applied
    has_styles = False
    for row in range(2, 12):  # Check rows 2-11 for banded or alternate row styling
        if (sheet.cell(row=row, column=1).fill.fill_type !=
            sheet.cell(row=row-1, column=1).fill.fill_type):
            has_styles = True
            break
    checklist_data["Completed"].append("Yes" if has_styles else "No")

    # 6. Check if headers in row 1 are bold and have a background color
    headers_bold = all(sheet.cell(row=1, column=col).font.bold for col in range(1, 8))
    headers_colored = all(sheet.cell(row=1, column=col).fill.fill_type for col in range(1, 8))
    checklist_data["Completed"].append("Yes" if headers_bold and headers_colored else "No")

    # 7. Check if columns are properly aligned
    correct_alignment = True
    for row in range(2, 12):  # Data rows
        for col in range(1, 8):  # All 7 columns
            cell = sheet.cell(row=row, column=col)
            alignment = cell.alignment.horizontal
            
            # Define expected alignment based on column index
            if col in [2, 3, 5, 6, 7]:  # Left-aligned columns
                expected_alignments = ["left", "general"]
            else:  # Center-aligned columns (1 and 4)
                expected_alignments = ["center", "general"]
            
            # Check if alignment matches expected values
            if alignment not in expected_alignments:
                correct_alignment = False
                break
        if not correct_alignment:
            break
    
    checklist_data["Completed"].append("Yes" if correct_alignment else "No")


    # 8. Check if ChatGPT hyperlink is centered in row 13
    center_aligned = sheet['A13'].alignment.horizontal == 'center'
    checklist_data["Completed"].append("Yes" if center_aligned else "No")

    # 9. Check if cells A13:G13 are merged
    merged_in_row_13 = any("A13:G13" in str(range) for range in sheet.merged_cells.ranges)
    checklist_data["Completed"].append("Yes" if merged_in_row_13 else "No")

    # 10. Check if row 13 has a background color and meaningful hyperlink
    background_color_present = sheet['A13'].fill.fill_type is not None
    meaningful_hyperlink = sheet['A13'].value is not None
    checklist_data["Completed"].append("Yes" if background_color_present and meaningful_hyperlink else "No")

    return checklist_data
