from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl import load_workbook

def check_excel_final(workbook):
    checklist_data = {
        "Grading Criteria": [
            "Are the worksheet names 'Workplace Productivity' and 'Department Distribution'?",
            "Does 'Workplace Productivity' have the required column headers?",
            "Does 'Workplace Productivity' have a summary row for company averages?",
            "Are the required charts present in 'Workplace Productivity'?",
            "Does 'Department Distribution' contain a table and pie chart?",
            "Is the data in 'Workplace Productivity' complete?",
            "Are the averages in the summary row accurate?",
            "Is the table in 'Workplace Productivity' logically sorted?",
            "Are the formatting and alignment consistent?",
            "Is the 'Training Requirements' column color-coded?"
        ],
        "Completed": []
    }

    # Check worksheet names
    required_sheets = ['Workplace Productivity', 'Department Distribution']
    sheet_names_correct = all(sheet in workbook.sheetnames for sheet in required_sheets)
    checklist_data["Completed"].append("Yes" if sheet_names_correct else "No")

    if not sheet_names_correct:
        return checklist_data  # Stop further checks if worksheets are missing

    wp_sheet = workbook['Workplace Productivity']
    dd_sheet = workbook['Department Distribution']

    # Check column headers
    expected_headers = ["Employee ID", "Department", "Digital Skills Score (1-10)", "Productivity Rating (1-5)",
                        "Hours of Training Completed", "Use of Productivity Software (hours/week)",
                        "Reported Weekly Output (Tasks Completed)", "Years at Company", "Age",
                        "Remote Work Percentage (%)", "Training Requirements"]
    headers = [wp_sheet.cell(row=1, column=i).value for i in range(1, len(expected_headers) + 1)]
    headers_match = all(a == b for a, b in zip(headers, expected_headers))
    checklist_data["Completed"].append("Yes" if headers_match else "No")


    # Validate that 'Company Averages' is in merged cells A17:B17
    merged_cell_label = wp_sheet['A17'].value
    summary_row_correct = (
        isinstance(merged_cell_label, str) and
        merged_cell_label.strip().lower() == 'company averages'
    )
    checklist_data["Completed"].append("Yes" if summary_row_correct else "No")


    # Check for charts in 'Workplace Productivity'
    wp_charts = [chart for chart in wp_sheet._charts]

    def check_chart_title(chart, expected_title):
        """Check chart titles in a case-insensitive manner."""
        if hasattr(chart, 'title') and chart.title is not None:
            if isinstance(chart.title, str):
                return chart.title.strip().lower() == expected_title.strip().lower()
            elif hasattr(chart.title, 'tx') and hasattr(chart.title.tx, 'rich'):
                for p in chart.title.tx.rich.paragraphs:
                    for run in p.r:
                        if hasattr(run, 't') and run.t.strip().lower() == expected_title.strip().lower():
                            return True
        return False

    has_digital_skills_chart = any(check_chart_title(chart, "Digital Skills Scores by Department") for chart in wp_charts)
    has_training_output_chart = any(check_chart_title(chart, "Hours of Training Completed and Reported Weekly Output") for chart in wp_charts)
    checklist_data["Completed"].append("Yes" if has_digital_skills_chart and has_training_output_chart else "No")

    # Check 'Department Distribution' table and chart
    table_correct = (dd_sheet.cell(row=1, column=1).value == 'Department' and
                     dd_sheet.cell(row=1, column=2).value == 'Number of Employees')
    pie_chart_correct = any(check_chart_title(chart, "Department Distribution") for chart in dd_sheet._charts)
    checklist_data["Completed"].append("Yes" if table_correct and pie_chart_correct else "No")

    # Check data completeness
    data_complete = all(all(wp_sheet.cell(row=row, column=col).value is not None for col in range(1, 12)) for row in range(2, 17))
    checklist_data["Completed"].append("Yes" if data_complete else "No")

    # Validate the presence of formulas in C17:J17
    formulas_present = True
    
    for col in range(3, 11):  # Columns C (3) to J (10)
        cell = wp_sheet.cell(row=17, column=col)
        if cell.data_type != 'f':  # 'f' indicates the cell contains a formula
            formulas_present = False
            print(f"Missing formula in Column {col}, Row 17")
    
    checklist_data["Completed"].append("Yes" if formulas_present else "No")

    # Check sorting
    sorted_correctly = True
    for i in range(2, 16):
        current_value = wp_sheet.cell(row=i, column=1).value
        next_value = wp_sheet.cell(row=i + 1, column=1).value
        if isinstance(current_value, (int, float)) and isinstance(next_value, (int, float)):
            if current_value > next_value:
                sorted_correctly = False
                break
    checklist_data["Completed"].append("Yes" if sorted_correctly else "No")

    # Check formatting and alignment
    consistent_formatting = all(
        wp_sheet.cell(row=1, column=col).alignment.horizontal == 'center' and
        wp_sheet.cell(row=1, column=col).font.bold for col in range(1, 12)
    )
    checklist_data["Completed"].append("Yes" if consistent_formatting else "No")

    # Check color-coded training requirements
    color_coded = any(
        wp_sheet.cell(row=row, column=11).fill is not None for row in range(2, 17)
    )
    checklist_data["Completed"].append("Yes" if color_coded else "No")

    return checklist_data

# Example usage:
# workbook = load_workbook('Final Project Example.xlsx')
# results = check_excel_final(workbook)
# print(results)
