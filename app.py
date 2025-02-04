import streamlit as st
from utils.display import display_results
from checkers.excel.excel_1 import check_excel_1
from checkers.excel.excel_2 import check_excel_2
from checkers.excel.excel_3 import check_excel_3
from checkers.excel.excel_final import check_excel_final
from checkers.word.word_1 import check_word_1
from openpyxl import load_workbook
from docx import Document
from pptx import Presentation

st.title("Assignment Checker")

# Excel Section
st.header("Excel Assignments")
excel_1_file = st.file_uploader("Upload Excel_1", type=["xlsx"], key="excel_1")
excel_2_file = st.file_uploader("Upload Excel_2", type=["xlsx"], key="excel_2")
excel_3_file = st.file_uploader("Upload Excel_3", type=["xlsx"], key="excel_3")
excel_final_file = st.file_uploader("Upload Excel_Final", type=["xlsx"], key="excel_final")

# Word Section
st.header("Word Assignments")
word_1_file = st.file_uploader("Upload Word_1", type=["docx"], key="word_1")

if word_1_file is not None:
    doc = Document(word_1_file)
    results = check_word_1(doc)

    # Display results
    for criteria, status in zip(results["Grading Criteria"], results["Completed"]):
        st.write(f"{criteria}: {status}")

# PowerPoint Section
st.header("PowerPoint Assignments")
ppt_1_file = st.file_uploader("Upload PowerPoint_1", type=["pptx"], key="ppt_1")

# Excel Checkers
if excel_1_file:
    try:
        workbook = load_workbook(excel_1_file)
        checklist_data = check_excel_1(workbook)
        st.subheader("Excel Assignment 1 Results")
        display_results(checklist_data)
    except Exception as e:
        st.error(f"An error occurred with Excel Assignment 1: {str(e)}")

if excel_2_file:
    try:
        workbook = load_workbook(excel_2_file)
        checklist_data = check_excel_2(workbook)
        st.subheader("Excel Assignment 2 Results")
        display_results(checklist_data)
    except Exception as e:
        st.error(f"An error occurred with Excel Assignment 2: {str(e)}")

if excel_3_file:
    try:
        workbook = load_workbook(excel_3_file)
        checklist_data = check_excel_3(workbook)
        st.subheader("Excel Assignment 3 Results")
        display_results(checklist_data)
    except Exception as e:
        st.error(f"An error occurred with Excel Assignment 3: {str(e)}")

if excel_final_file:
    try:
        workbook = load_workbook(excel_final_file)
        checklist_data = check_excel_final(workbook)  # Pass workbook directly
        st.subheader("Excel Final Assignment Results")
        display_results(checklist_data)
    except Exception as e:
        st.error(f"An error occurred with Excel Final Assignment: {str(e)}")

# Word Checker
if word_1_file:
    try:
        doc = Document(word_1_file)
        from checkers.word.word_1 import check_word_1
        checklist_data = check_word_1(doc)
        st.subheader("Word Assignment 1 Results")
        display_results(checklist_data)
    except Exception as e:
        st.error(f"An error occurred with Word Assignment 1: {str(e)}")

# PowerPoint Checker
if ppt_1_file:
    try:
        prs = Presentation(ppt_1_file)
        from checkers.powerpoint.ppt_1 import check_ppt_1
        checklist_data = check_ppt_1(prs)
        st.subheader("PowerPoint Assignment 1 Results")
        display_results(checklist_data)
    except Exception as e:
        st.error(f"An error occurred with PowerPoint Assignment 1: {str(e)}")
